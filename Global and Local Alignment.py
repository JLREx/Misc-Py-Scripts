# This code is written for usage in subject TIB2341 Introduction to Bioinformatics. 
# You might need to install "windnd" and "openpyxl" to run this code. 
# This is not a professional code by any means, it is just a simple program to fulfill my assignment requirement. 
# Also, there are still a few things here and there that can be further develop for performance, design, and user experience improvement. 

# See https://openpyxl.readthedocs.io/en/stable/ for documentation of the openpyxl library

import numpy as np
import windnd
import re

import tkinter as tk
from tkinter import ttk
from ttkthemes import ThemedTk
from tkinter import filedialog
from tkinter.messagebox import showinfo

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.utils import rows_from_range

def dragged(files):
    if (len(files) > 2):
        exceed_message = "This program only accepts up to two files at a time."
        showinfo("Too many files", exceed_message)
        return

    fileList = []
    for file in files:
        file = file.decode('UTF-8')
        fileList.append(file)

    fileTypeChecking(fileList)

def selectFile():
    files = filedialog.askopenfilenames(initialdir='/', title='Select files', 
                                        filetypes=(("Fasta Files", "*.fasta"), ("Text Files", "*.txt"), 
                                                   ("All Files", "*.*")))
    fileTypeChecking(files)
    
def fileTypeChecking(files):
    nonTxtFlag = False
    for file in files:
        file_splitted = re.split("\.", file)
        if (file_splitted[:][-1] != "txt" and file_splitted[:][-1] != "fasta"):
            error_message = "Found unsupported file format loading " + file
            showinfo("File format not supported", error_message)
            nonTxtFlag = True
            break

    if not nonTxtFlag: 
        run(files)

def run(files):
    if len(files) < 2:
        with open(files[0]) as sequenceFile :
            sequenceNames = []
            sequences = []
            foundOnce = False
            foundTwice = False
            for line in sequenceFile :
                if '>' in line and foundTwice:
                    break
                elif '>' in line and foundOnce:
                    sequenceNames.append(line[1:].rstrip())
                    sequences.append(sequence)
                    sequence = ''
                    foundTwice = True
                elif '>' in line:
                    sequenceNames.append(line[1:].rstrip())
                    sequence = ''
                    foundOnce = True
                else:
                    sequence = sequence + line.rstrip()
            sequences.append(sequence)

        firstSequence, secondSequence = sequences[0], sequences[1]
        firstSequenceName, secondSequenceName = sequenceNames[0].split(), sequenceNames[1].split()
        firstSequenceName, secondSequenceName = firstSequenceName[0], secondSequenceName[0]
    
    elif len(files) == 2:
        with open(files[0]) as firstSequenceFile :
            foundOnce = False
            for line in firstSequenceFile :
                if '>' in line and foundOnce:
                    break
                elif '>' in line:
                    firstSequenceName = line[1:].rstrip()
                    firstSequence = ''
                    foundOnce = True
                else:
                    firstSequence = firstSequence + line.rstrip()

        with open(files[1]) as secondSequenceFile :
            foundOnce = False
            for line in secondSequenceFile :
                if '>' in line and foundOnce:
                    break
                elif '>' in line:
                    secondSequenceName = line[1:].rstrip()
                    secondSequence = ''
                    foundOnce = True
                else:
                    secondSequence = secondSequence + line.rstrip()
    
    txt_first_sequence.delete(1.0, 'end')
    txt_second_sequence.delete(1.0, 'end')
    txt_first_sequence.insert(1.0, firstSequence)
    txt_second_sequence.insert(1.0, secondSequence)
    
    button_val = radio_indicator.get()
    if (button_val == 1):
        check_global(firstSequence, secondSequence, firstSequenceName, secondSequenceName)
    else:
        check_local(firstSequence, secondSequence, firstSequenceName, secondSequenceName)
        
def getLine():
    firstSequence = txt_first_sequence.get('1.0', 'end').strip()
    secondSequence = txt_second_sequence.get('1.0', 'end').strip()
    
    firstSequenceName, secondSequenceName = firstSequence[:5], secondSequence[:5]
    
    button_val = radio_indicator.get()
    if (button_val == 1):
        check_global(firstSequence, secondSequence, firstSequenceName, secondSequenceName)
    else:
        check_local(firstSequence, secondSequence, firstSequenceName, secondSequenceName)
    
def base_matrix_global(gapPenalty, firstSequence, secondSequence):
    matrix = np.zeros((len(firstSequence) + 1, len(secondSequence) + 1))
    for i in range(len(firstSequence) + 1):
        matrix[i][0] = i * gapPenalty
    for j in range(len(secondSequence) + 1):
        matrix[0][j] = j * gapPenalty
    return matrix

def base_matrix_local(firstSequence, secondSequence):
    matrix = np.zeros((len(firstSequence) + 1, len(secondSequence) + 1))
    for i in range(len(firstSequence) + 1):
        matrix[i][0] = i * 0
    for j in range(len(secondSequence) + 1):
        matrix[0][j] = j * 0
    return matrix

def fill_matrix_global(matrix, matchReward, mismatchPenalty, gapPenalty, firstSequence, secondSequence):
    for i in range(1, len(firstSequence)+1):
        for j in range(1, len(secondSequence)+1):
            if firstSequence[i-1] == secondSequence[j-1]:
                matrix[i][j] = max(matrix[i-1][j-1] + matchReward, 
                                  matrix[i-1][j] + gapPenalty, 
                                  matrix[i][j-1] + gapPenalty)
            else:
                matrix[i][j] = max(matrix[i-1][j-1] + mismatchPenalty, 
                                  matrix[i-1][j] + gapPenalty, 
                                  matrix[i][j-1] + gapPenalty)
    return matrix

def fill_matrix_local(matrix, matchReward, mismatchPenalty, gapPenalty, firstSequence, secondSequence):
    for i in range(1, len(firstSequence)+1):
        for j in range(1, len(secondSequence)+1):
            if firstSequence[i-1] == secondSequence[j-1]:
                matrix[i][j] = max(matrix[i-1][j-1] + matchReward, 
                                  matrix[i-1][j] + gapPenalty, 
                                  matrix[i][j-1] + gapPenalty)
            else:
                matrix[i][j] = max(matrix[i-1][j-1] + mismatchPenalty, 
                                  matrix[i-1][j] + gapPenalty, 
                                  matrix[i][j-1] + gapPenalty)
            if matrix[i][j] < 0:
                matrix[i][j] = 0 
    return matrix

def alignment_check_global(matrix, matchReward, mismatchPenalty, gapPenalty, firstSequence, secondSequence):
    i, j = len(firstSequence), len(secondSequence)
    matchedFirstSequenceList, matchedSecondSequenceList, signsList = [''], [''], ['']
    firstSequenceFinal, secondSequenceFinal, signsFinal, tracesFinal = [], [], [], []
    traceCoordinatesList = [[(i, j)]]
    alternativeCoordinatesList = []
    count = 1
    totalCount = 0
    matchedFirstSequence, matchedSecondSequence, signs = '', '', ''
    traceCoordinates = [(i, j)]
    
    while (count > 0):
        matchedFirstSequence = matchedFirstSequenceList[totalCount]
        matchedSecondSequence = matchedSecondSequenceList[totalCount]
        signs = signsList[totalCount]
        traceCoordinates = traceCoordinatesList[totalCount]
        totalCount += 1
        
        while (i > 0 and j > 0):
            if matrix[i-1][j-1] + matchReward == matrix[i][j] and firstSequence[i-1] == secondSequence[j-1]:
                if matrix[i-1][j] + gapPenalty == matrix[i][j] and matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i-1, j))
                    matchedFirstSequenceList.append(firstSequence[i-1] + matchedFirstSequence)
                    matchedSecondSequenceList.append('-' + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i-1, j))
                    traceCoordinatesList.append(newCoordinates)
                    
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 2

                elif matrix[i-1][j] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i-1, j))
                    matchedFirstSequenceList.append(firstSequence[i-1] + matchedFirstSequence)
                    matchedSecondSequenceList.append('-' + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i-1, j))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1

                elif matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1

                matchedFirstSequence = firstSequence[i-1] + matchedFirstSequence
                matchedSecondSequence = secondSequence[j-1] + matchedSecondSequence
                signs = '|' + signs
                i -= 1
                j -= 1

            elif matrix[i-1][j-1] + mismatchPenalty == matrix[i][j]:
                if matrix[i-1][j] + gapPenalty == matrix[i][j] and matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i-1, j))
                    matchedFirstSequenceList.append(firstSequence[i-1] + matchedFirstSequence)
                    matchedSecondSequenceList.append('-' + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i-1, j))
                    traceCoordinatesList.append(newCoordinates)
                    
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 2

                elif matrix[i-1][j] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i-1, j))
                    matchedFirstSequenceList.append(firstSequence[i-1] + matchedFirstSequence)
                    matchedSecondSequenceList.append('-' + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i-1, j))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1

                elif matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1

                matchedFirstSequence = firstSequence[i-1] + matchedFirstSequence
                matchedSecondSequence = secondSequence[j-1] + matchedSecondSequence
                signs = ' ' + signs
                i -= 1
                j -= 1

            else:
                if matrix[i-1][j] + gapPenalty == matrix[i][j] and matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1
                    
                    matchedFirstSequence = firstSequence[i-1] + matchedFirstSequence
                    matchedSecondSequence = '-' + matchedSecondSequence
                    signs = ' ' + signs
                    i -= 1

                elif matrix[i-1][j] + gapPenalty == matrix[i][j]:
                    matchedFirstSequence = firstSequence[i-1] + matchedFirstSequence
                    matchedSecondSequence = '-' + matchedSecondSequence
                    signs = ' ' + signs
                    i -= 1

                elif matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    matchedFirstSequence = '-' + matchedFirstSequence
                    matchedSecondSequence = secondSequence[j-1] + matchedSecondSequence
                    signs = ' ' + signs
                    j -= 1
            traceCoordinates.append((i, j))

        if i > 0:
            for remaining in range(i, 0, -1):
                matchedFirstSequence = firstSequence[remaining-1] + matchedFirstSequence
                matchedSecondSequence = '-' + matchedSecondSequence
                signs = ' ' + signs
                traceCoordinates.append((i, j))

        if j > 0:
            for remaining in range(j, 0, -1):
                matchedFirstSequence = '-' + matchedFirstSequence
                matchedSecondSequence = secondSequence[remaining-1] + matchedSecondSequence
                signs = ' ' + signs
                traceCoordinates.append((i, j))
        
        firstSequenceFinal.append(matchedFirstSequence)
        secondSequenceFinal.append(matchedSecondSequence)
        signsFinal.append(signs)
        tracesFinal.append(traceCoordinates)
        
        count -= 1
        if count > 0:
            i, j = alternativeCoordinatesList[totalCount-1]
    
    return firstSequenceFinal, secondSequenceFinal, signsFinal, tracesFinal

def alignment_check_local(matrix, matchReward, mismatchPenalty, gapPenalty, firstSequence, secondSequence):
    maxValue = np.amax(matrix)
    result = np.where(matrix == np.amax(matrix))
    i, j = result[0][0], result[1][0]
    
    matchedFirstSequenceList, matchedSecondSequenceList, signsList = [''], [''], ['']
    firstSequenceFinal, secondSequenceFinal, signsFinal, tracesFinal = [], [], [], []
    traceCoordinatesList = [[(i, j)]]
    alternativeCoordinatesList = []
    count = 1
    totalCount = 0
    matchedFirstSequence, matchedSecondSequence, signs = '', '', ''
    traceCoordinates = [(i, j)]
    
    while (count > 0):
        matchedFirstSequence = matchedFirstSequenceList[totalCount]
        matchedSecondSequence = matchedSecondSequenceList[totalCount]
        signs = signsList[totalCount]
        traceCoordinates = traceCoordinatesList[totalCount]
        totalCount += 1
        
        while (matrix[i][j]!=0): 
            if matrix[i-1][j-1] + matchReward == matrix[i][j] and firstSequence[i-1] == secondSequence[j-1]:
                if matrix[i-1][j] + gapPenalty == matrix[i][j] and matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i-1, j))
                    matchedFirstSequenceList.append(firstSequence[i-1] + matchedFirstSequence)
                    matchedSecondSequenceList.append('-' + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i-1, j))
                    traceCoordinatesList.append(newCoordinates)
                    
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 2

                elif matrix[i-1][j] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i-1, j))
                    matchedFirstSequenceList.append(firstSequence[i-1] + matchedFirstSequence)
                    matchedSecondSequenceList.append('-' + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i-1, j))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1

                elif matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1

                matchedFirstSequence = firstSequence[i-1] + matchedFirstSequence
                matchedSecondSequence = secondSequence[j-1] + matchedSecondSequence
                signs = '|' + signs
                i -= 1
                j -= 1

            elif matrix[i-1][j-1] + mismatchPenalty == matrix[i][j]:
                if matrix[i-1][j] + gapPenalty == matrix[i][j] and matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i-1, j))
                    matchedFirstSequenceList.append(firstSequence[i-1] + matchedFirstSequence)
                    matchedSecondSequenceList.append('-' + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i-1, j))
                    traceCoordinatesList.append(newCoordinates)
                    
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 2

                elif matrix[i-1][j] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i-1, j))
                    matchedFirstSequenceList.append(firstSequence[i-1] + matchedFirstSequence)
                    matchedSecondSequenceList.append('-' + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i-1, j))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1

                elif matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1

                matchedFirstSequence = firstSequence[i-1] + matchedFirstSequence
                matchedSecondSequence = secondSequence[j-1] + matchedSecondSequence
                signs = ' ' + signs
                i -= 1
                j -= 1

            else:
                if matrix[i-1][j] + gapPenalty == matrix[i][j] and matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    alternativeCoordinatesList.append((i, j-1))
                    matchedFirstSequenceList.append('-' + matchedFirstSequence)
                    matchedSecondSequenceList.append(secondSequence[j-1] + matchedSecondSequence)
                    signsList.append(' ' + signs)
                    newCoordinates = traceCoordinates.copy()
                    newCoordinates.append((i, j-1))
                    traceCoordinatesList.append(newCoordinates)
                    count += 1
                    
                    matchedFirstSequence = firstSequence[i-1] + matchedFirstSequence
                    matchedSecondSequence = '-' + matchedSecondSequence
                    signs = ' ' + signs
                    i -= 1

                elif matrix[i-1][j] + gapPenalty == matrix[i][j]:
                    matchedFirstSequence = firstSequence[i-1] + matchedFirstSequence
                    matchedSecondSequence = '-' + matchedSecondSequence
                    signs = ' ' + signs
                    i -= 1

                elif matrix[i][j-1] + gapPenalty == matrix[i][j]:
                    matchedFirstSequence = '-' + matchedFirstSequence
                    matchedSecondSequence = secondSequence[j-1] + matchedSecondSequence
                    signs = ' ' + signs
                    j -= 1
            traceCoordinates.append((i, j))

        firstSequenceFinal.append(matchedFirstSequence)
        secondSequenceFinal.append(matchedSecondSequence)
        signsFinal.append(signs)
        tracesFinal.append(traceCoordinates)
        
        count -= 1
        if count > 0:
            i, j = alternativeCoordinatesList[totalCount-1]
    
    return firstSequenceFinal, secondSequenceFinal, signsFinal, tracesFinal

def check_global(firstSequence, secondSequence, firstSequenceName, secondSequenceName):
    matchReward = int(txt_match.get())
    mismatchPenalty = int(txt_mismatch.get())
    gapPenalty = int(txt_gap.get())
    
    firstSequence, secondSequence = firstSequence.upper(), secondSequence.upper()
    
    matrix = base_matrix_global(gapPenalty, firstSequence, secondSequence)
    matrix = fill_matrix_global(matrix, matchReward, mismatchPenalty, gapPenalty, firstSequence, secondSequence)
    matchedFirstSequenceList, matchedSecondSequenceList, signsList, traceCoordinatesList = alignment_check_global(matrix, matchReward, mismatchPenalty, gapPenalty, firstSequence, secondSequence)
    
    scoreList = []
    for i in range(len(matchedFirstSequenceList)):
        gapCount = matchedFirstSequenceList[i].count('-') + matchedSecondSequenceList[i].count('-')
        mismatchCount = signsList[i].count(' ')
        matchCount = signsList[i].count('|')
        correctedMismatchCount = mismatchCount - gapCount
        score = gapCount * gapPenalty + correctedMismatchCount * mismatchPenalty + matchCount * matchReward
        scoreList.append(score)

    txt_alignment.config(state='normal')
    txt_alignment.delete(1.0, 'end')
    txt_alignment.insert(1.0, matchedFirstSequenceList[0] + "\n")
    txt_alignment.insert(2.0, signsList[0] + "\n")
    txt_alignment.insert(3.0, matchedSecondSequenceList[0])
    txt_alignment.config(state='disabled')
    
    outputCSV(firstSequence, secondSequence, matchedFirstSequenceList, 
              signsList, matchedSecondSequenceList, matrix, matchReward, mismatchPenalty, gapPenalty, scoreList, 
              traceCoordinatesList, firstSequenceName, secondSequenceName)
    
def check_local(firstSequence, secondSequence, firstSequenceName, secondSequenceName):
    matchReward = int(txt_match.get())
    mismatchPenalty = int(txt_mismatch.get())
    gapPenalty = int(txt_gap.get())
    
    firstSequence, secondSequence = firstSequence.upper(), secondSequence.upper()
    
    matrix = base_matrix_local(firstSequence, secondSequence)
    matrix = fill_matrix_local(matrix, matchReward, mismatchPenalty, gapPenalty, firstSequence, secondSequence)
    matchedFirstSequenceList, matchedSecondSequenceList, signsList, traceCoordinatesList = alignment_check_local(matrix, matchReward, mismatchPenalty, gapPenalty, firstSequence, secondSequence)
    
    scoreList = []
    for i in range(len(matchedFirstSequenceList)):
        gapCount = matchedFirstSequenceList[i].count('-') + matchedSecondSequenceList[i].count('-')
        mismatchCount = signsList[i].count(' ')
        matchCount = signsList[i].count('|')
        correctedMismatchCount = mismatchCount - gapCount
        score = gapCount * gapPenalty + correctedMismatchCount * mismatchPenalty + matchCount * matchReward
        scoreList.append(score)
    
    txt_alignment.config(state='normal')
    txt_alignment.delete(1.0, 'end')
    txt_alignment.insert(1.0, matchedFirstSequenceList[0] + "\n")
    txt_alignment.insert(2.0, signsList[0] + "\n")
    txt_alignment.insert(3.0, matchedSecondSequenceList[0])
    txt_alignment.config(state='disabled')
    
    outputCSV(firstSequence, secondSequence, matchedFirstSequenceList, 
              signsList, matchedSecondSequenceList, matrix, matchReward, mismatchPenalty, gapPenalty, scoreList, 
              traceCoordinatesList, firstSequenceName, secondSequenceName)
    
def outputCSV(firstSequence, secondSequence, matchedFirstSequence, 
              signs, matchedSecondSequence, matrix, matchReward, mismatchPenalty, gapPenalty, scoreList, 
              traceCoordinatesList, firstSequenceName, secondSequenceName):
    wb = Workbook()
    ws = wb.active
    ws.title = "Best Alignment"
    yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    centerAlign = Alignment(horizontal = 'center', vertical='center', wrap_text=True)    
    button_val = radio_indicator.get()

    for i in range(0, len(traceCoordinatesList)):
        if (button_val == 1):
            ws['B1'].value = "Global Alignment"
        else:
            ws['B1'].value = "Local Alignment"

        ws['B2'].value = "Sequence 1"
        ws['B3'].value = "Sequence 2"
        ws.merge_cells('B2:C2')
        ws.merge_cells('B3:C3')
        ws['D2'].value = firstSequence
        ws['D3'].value = secondSequence
        ws['B5'].value = "Gap"
        ws['C5'].value = gapPenalty
        ws['B6'].value = "Score"
        ws['B6'].fill = yellowFill
        ws['B6'].font = Font(bold=True)
        ws['C6'].value = scoreList[i]
        ws['C6'].fill = yellowFill
        ws['C6'].font = Font(bold=True)
        ws['E5'].value = "Match"
        ws['F5'].value = matchReward
        ws['H5'].value = "Mismatch"
        ws['I5'].value = mismatchPenalty

        for column in range(2, len(matchedFirstSequence[i])+2):
            ws.cell(column=column, row=8, value=matchedFirstSequence[i][column-2])
            ws.cell(column=column, row=8).alignment = centerAlign
            ws.cell(column=column, row=9, value=signs[i][column-2])
            ws.cell(column=column, row=9).alignment = centerAlign
            ws.cell(column=column, row=10, value=matchedSecondSequence[i][column-2])
            ws.cell(column=column, row=10).alignment = centerAlign

        for column in range(4, len(secondSequence)+4):
            ws.cell(column=column, row=12, value=secondSequence[column-4])
            ws.cell(column=column, row=12).font = Font(bold=True)
            ws.cell(column=column, row=12).alignment = centerAlign
        for row in range(14, len(firstSequence)+14):
            ws.cell(column=2, row=row, value=firstSequence[row-14])
            ws.cell(column=2, row=row).font = Font(bold=True)
            ws.cell(column=2, row=row).alignment = centerAlign

        for row in range(len(matrix)):
            for column in range(len(matrix[row])):
                ws.cell(column=column+3, row=row+13, value=matrix[row][column])
                ws.cell(row+13, column+3).alignment = centerAlign
                if (row, column) in traceCoordinatesList[i]:
                    ws.cell(row+13, column+3).fill = yellowFill
                    ws.cell(row+13, column+3).font = Font(bold=True)
        
        ws = wb.create_sheet(title='Alignment ' + str(i+1))
    
    wb.remove(ws)
    name = firstSequenceName + " & " + secondSequenceName
    wb.save(name + ".xlsx")

class SinglePage(tk.Frame):
    def __init__(self, *args, **kwargs):
        tk.Frame.__init__(self, *args, **kwargs)
        frm_page = tk.Frame(self).pack()
        
        style = ttk.Style()
        style.configure('TButton', font=("Verdana", 10))
        
        global radio_indicator
        global txt_first_sequence, txt_second_sequence, txt_alignment
        global txt_match, txt_mismatch, txt_gap
        
        lbl_heading = ttk.Label(frm_page, text='Sequence Alignment', padding=(0, 10, 0, 10), 
                               font=("Bahnschrift SemiBold", 18)).grid(column=1)
        lbl_empty = ttk.Label(frm_page).grid(row=2, columnspan=3, rowspan=2, pady=15)
        
        radio_indicator = tk.IntVar(value=1)
        btn_radio_global = ttk.Radiobutton(frm_page, text='Global Alignment', 
                                           value=1, variable=radio_indicator, 
                                           width=20).place(x=135, y=60)
        btn_radio_local = ttk.Radiobutton(frm_page, text='Local Alignment', 
                                          value=2, variable=radio_indicator, 
                                          width=20).place(x=320, y=60)
        
        txt_gap = tk.IntVar(value=-1)
        txt_match = tk.IntVar(value=1)
        txt_mismatch = tk.IntVar(value=0)
        
        lbl_gap = ttk.Label(frm_page, text='Gap').grid(row=4, column=0)
        lbl_match = ttk.Label(frm_page, text='Match').grid(row=4, column=1)
        lbl_mismatch = ttk.Label(frm_page, text='Mismatch').grid(row=4, column=2)
        
        ent_gap = ttk.Entry(frm_page, textvariable=txt_gap)
        ent_gap.grid(row=5, column=0)
        ent_match = ttk.Entry(frm_page, textvariable=txt_match)
        ent_match.grid(row=5, column=1)
        ent_mismatch = ttk.Entry(frm_page, textvariable=txt_mismatch)
        ent_mismatch.grid(row=5, column=2)
                
        lbl_first_sequence = ttk.Label(frm_page, text='First Sequence', 
                                       padding=(0, 10, 0, 5)).grid(row=7, columnspan=3)
        txt_first_sequence = tk.Text(frm_page, height=4, padx=10, pady=3, font=('Consolas', 10))
        txt_first_sequence.grid(row=8, columnspan=3)
        
        lbl_second_sequence = ttk.Label(frm_page, text='Second Sequence', 
                                        padding=(0, 10, 0, 5)).grid(row=9, columnspan=3)
        txt_second_sequence = tk.Text(frm_page, height=4, padx=10, pady=3, font=('Consolas', 10))
        txt_second_sequence.grid(row=10, columnspan=3)
        
        lbl_result = ttk.Label(frm_page, text='Best Result', 
                               padding=(0, 10, 0, 5)).grid(row=11, columnspan=3)
        txt_alignment = tk.Text(frm_page, state='disabled', height=3, padx=10, pady=3, font=('Consolas', 10))
        txt_alignment.grid(row=12, columnspan=3)
        
        btn_select = ttk.Button(frm_page, text='Select Files', width=25, 
                                command=selectFile).place(x=40, y=465)
        btn_check = ttk.Button(frm_page, text='Check Alignment', width=25, 
                               command=getLine).place(x=320, y=465)

if __name__ == "__main__":
    root = ThemedTk(theme='breeze')
    root.title("Sequence Alignment")
    root.geometry('580x520')
    root.resizable(0, 0)
    windnd.hook_dropfiles(root, func=dragged)
    page = SinglePage(root)
    root.mainloop()